import openpyxl
from csv_to_series_entries import csv_to_series_entries
from functions.sortFuncs import get_teams_carNums

csv_file = './entries_23.csv'
event_name = 'Sonoma Raceway'


def create_manager_signin(series_entries, event):
    wb = openpyxl.load_workbook('./signin_templates/manager_signin.xlsx')

    for series, entries in series_entries.items():
        sheet = wb[series]

        team_car_dict = get_teams_carNums(entries)

        current_row = 8

        for team, car_nums in team_car_dict.items():
            sheet.cell(row=current_row, column=1, value=team)

            sheet.cell(row=current_row, column=2, value=', '.join(car_nums))

            current_row += 1

    wb.save(f'manager_signin/{event}.xlsx')


create_manager_signin(csv_to_series_entries(csv_file, event_name), event_name)
