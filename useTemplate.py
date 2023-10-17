import openpyxl
from csv_converter import getWCEntries

template = '../signin_templates/driver_master.xlsx'
csv_file = './RA_entries.csv'

# {'Driver Designation': 'Pro - Am', 'Team Name': 'RealTime', 'number': '43', 'event': 'Road America', 'series': 'GTWCA', 'driver2': 'Adam Christodoulou', 'driver1': 'Anthony Bartone'}

header_mapping = {
    'number': 'Car #',
    'Team Name': 'Team',
    'driver1': 'Driver 1',
    'driver2': 'Driver 2',
    'Signature': 'Signature',
    'Signature2': 'Signature2'
}


def fill_excel_template(data, template):
    wb = openpyxl.load_workbook(template)
    event = data[0]['event']
    series = data[0]['series']
    sheet = wb[series]

    for row_i, row_data in enumerate(data, start=8):
        for col_i, key in enumerate(row_data.keys(), start=1):

            if key in header_mapping:

                mapped_header = header_mapping[key]
                # Find the column index for the mapped header
                for col in range(1, sheet.max_column + 1):
                    if sheet.cell(row=7, column=col).value == mapped_header:
                        col_i = col
                        break

                sheet.cell(row=row_i, column=col_i).value = row_data[key]
    wb.save(f'./Driver_Sign_in/{series}_{event}.xlsx')


fill_excel_template(getWCEntries(csv_file), template)
