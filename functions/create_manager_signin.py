import openpyxl
from functions.utility import get_all_teams
from functions.sortFuncs import get_teams_carNums


def create_manager_signin(series_entries, event):
    wb = openpyxl.load_workbook('./signin_templates/manager_signin.xlsx')

    for series, entries in series_entries.items():
        sheet = wb[series]

        team_car_arr = get_teams_carNums(entries)

        current_row = 8

        for team in team_car_arr:
            sheet.cell(row=current_row, column=1, value=team[0])

            sheet.cell(row=current_row, column=2, value=', '.join(team[1]))

            current_row += 1

    # handle event logistics
    sheet = wb['Event_log']

    all_entries = get_all_teams(series_entries)

    current = 8

    for team_name in all_entries:
        sheet.cell(row=current, column=1, value=team_name)

        current += 1

    wb.save(f'manager_signin/{event}.xlsx')
