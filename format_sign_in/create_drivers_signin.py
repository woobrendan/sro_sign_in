import openpyxl
import os


def create_drivers_signin(series_entries, event):
    wb = openpyxl.load_workbook('./templates/driver_master.xlsx')

    drivers_direct = "./sign_in_sheets/drivers/"

    if not os.path.exists(drivers_direct):
        os.makedirs(drivers_direct)

    for series, entries in series_entries.items():
        sheet = wb[series]

        current = 8

        for entry in entries:
            sheet.cell(row=current, column=1, value=entry['number'])
            sheet.cell(row=current, column=2, value=entry['team'])
            sheet.cell(row=current, column=3,
                       value=f'{entry["driver1firstName"]} {entry["driver1lastName"]}')

            if series == 'GTWCA' or series == 'PGT4A':
                sheet.cell(row=current, column=5,
                           value=f'{entry["driver2firstName"]} {entry["driver2lastName"]}')

            current += 1

    wb.save(f'./sign_in_sheets/drivers/{event}.xlsx')
