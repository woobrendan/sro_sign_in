import openpyxl
from format_sign_in.event_log import event_log
from utility.sortFuncs import get_teams_carNums
import os


def create_manager_signin(series_entries, event):
    wb = openpyxl.load_workbook('./templates/manager_signin.xlsx')
    manager_direct = "./sign_in_sheets/manager/"

    if not os.path.exists(manager_direct):
        os.makedirs(manager_direct)


    for series, entries in series_entries.items():
        sheet = wb[series]

        team_car_arr = get_teams_carNums(entries)

        current_row = 8

        for team in team_car_arr:
            sheet.cell(row=current_row, column=1, value=team[0])

            sheet.cell(row=current_row, column=2, value=', '.join(team[1]))

            current_row += 1

    # handle event logistics page
    event_log(wb, series_entries)

    wb.save(f'./sign_in_sheets/manager/{event}.xlsx')
