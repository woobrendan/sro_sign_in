import openpyxl
from csv_to_series_entries import csv_to_series_entries

csv_file = './RA_entries.csv'

# {'Driver Designation': 'Pro - Am', 'Team Name': 'RealTime', 'number': '43', 'event': 'Road America', 'series': 'GTWCA', 'driver2': 'Adam Christodoulou', 'driver1': 'Anthony Bartone'}

header_mapping = {
    'number': 'Car #',
    'Team Name': 'Team',
    'driver1': 'Driver 1',
    'driver2': 'Driver 2',
    'Signature': 'Signature',
    'Signature2': 'Signature2'
}


def create_drivers_signin(series_entries):
    wb = openpyxl.load_workbook('./signin_templates/driver_master.xlsx')
    event = series_entries['GTAM'][0]['event']

    # Loop through each series key and go through entries to complete data
    for series, entries in series_entries.items():
        sheet = wb[series]

        for row_i, entry in enumerate(entries, start=8):
            for col_i, key in enumerate(entry.keys(), start=1):

                if key in header_mapping:

                    mapped_header = header_mapping[key]
                    # Find the column index for the mapped header
                    for col in range(1, sheet.max_column + 1):
                        if sheet.cell(row=7, column=col).value == mapped_header:
                            col_i = col
                            break

                    sheet.cell(row=row_i, column=col_i).value = entry[key]
    wb.save(f'./Driver_Sign_in/{event}.xlsx')


create_drivers_signin(csv_to_series_entries(csv_file))
