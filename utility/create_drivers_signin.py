import openpyxl
from utility.csv_to_series_entries import csv_to_series_entries

# {'Driver Designation': 'Pro - Am', 'Team Name': 'RealTime', 'number': '43', 'event': 'Road America', 'series': 'GTWCA', 'driver2': 'Adam Christodoulou', 'driver1': 'Anthony Bartone'}


def create_drivers_signin(series_entries, event):
    wb = openpyxl.load_workbook('./templates/driver_master.xlsx')

    for series, entries in series_entries.items():
        sheet = wb[series]

        current = 8

        for entry in entries:
            sheet.cell(row=current, column=1, value=entry['number'])
            sheet.cell(row=current, column=2, value=entry['Team Name'])
            sheet.cell(row=current, column=3, value=entry['driver1'])

            if series == 'GTWCA' or series == 'PGT4A':
                sheet.cell(row=current, column=5, value=entry['driver2'])

            current += 1

    wb.save(f'./Driver_Sign_in/{event}.xlsx')
